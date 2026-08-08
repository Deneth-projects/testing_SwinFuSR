import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
import helpers

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_header(ws, headers, row=1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(12, length + 2)


def _write_summary_rows(ws, summary, whole_dataset_max, whole_dataset_max_percent, start_row=2):
    row_idx = start_row
    for label, stats in summary.items():
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=helpers.round4(stats["mean"]))
        ws.cell(row=row_idx, column=3, value=helpers.round4(stats["std"]))
        ws.cell(row=row_idx, column=4, value=helpers.round4(stats["min"]))
        ws.cell(row=row_idx, column=5, value=helpers.round4(stats["max"]))
        row_idx += 1

    ws.cell(row=row_idx, column=1, value="Whole-Dataset Maximum Error").font = Font(italic=True)
    ws.cell(row=row_idx, column=5, value=helpers.round4(whole_dataset_max))
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Whole-Dataset Maximum Error (%)").font = Font(italic=True)
    ws.cell(row=row_idx, column=5, value=helpers.round4(whole_dataset_max_percent))
    return row_idx + 1


def export_excel(results, summary, whole_dataset_max, whole_dataset_max_percent,
                  n_found, n_evaluated, n_skipped, log_path=None):
    if not config.SAVE_EXCEL:
        return None

    excel_folder = helpers.ensure_dir(os.path.join(config.OUTPUT_FOLDER, "Excel"))
    out_path = os.path.join(excel_folder, "evaluation_report.xlsx")

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Image Metrics"
    headers = ["Pair ID", "PSNR (dB)", "SSIM", "MS-SSIM", "LPIPS",
               "MAE", "MAE (%)", "RMSE", "RMSE (%)",
               "Maximum Error", "Maximum Error (%)"]
    _write_header(ws1, headers)
    for row_idx, r in enumerate(results, start=2):
        values = [
            r["pair_id"],
            helpers.round4(r["psnr"]),
            helpers.round4(r["ssim"]),
            helpers.round4(r["ms_ssim"]),
            helpers.round4(r["lpips"]) if r["lpips"] is not None else "N/A",
            helpers.round4(r["mae"]),
            helpers.round4(r["mae_percent"]),
            helpers.round4(r["rmse"]),
            helpers.round4(r["rmse_percent"]),
            helpers.round4(r["max_error"]),
            helpers.round4(r["max_error_percent"]),
        ]
        for col_idx, val in enumerate(values, start=1):
            ws1.cell(row=row_idx, column=col_idx, value=val)
    _autosize(ws1)

    ws2 = wb.create_sheet("Summary Statistics")
    _write_header(ws2, ["Metric", "Mean", "Std Dev", "Min", "Max"])
    _write_summary_rows(ws2, summary, whole_dataset_max, whole_dataset_max_percent)
    _autosize(ws2)

    ws3 = wb.create_sheet("Evaluation Information")
    info_rows = [
        ("Evaluation Date/Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Pairs Found", n_found),
        ("Pairs Evaluated", n_evaluated),
        ("Pairs Skipped", n_skipped),
        ("LR Folder", config.LR_FOLDER),
        ("RGB Folder", config.RGB_FOLDER),
        ("HR Ground Truth Folder", config.HR_GT_FOLDER),
        ("Predicted Folder", config.PRED_FOLDER),
        ("Output Folder", config.OUTPUT_FOLDER),
        ("Expected Image Size (W x H)", f"{config.INPUT_IMAGE_SIZE[0]} x {config.INPUT_IMAGE_SIZE[1]}"),
        ("Data Range", config.DATA_RANGE),
        ("LPIPS Network", config.LPIPS_NET),
        ("Log File", log_path or "-"),
    ]
    _write_header(ws3, ["Field", "Value"])
    for row_idx, (field, value) in enumerate(info_rows, start=2):
        ws3.cell(row=row_idx, column=1, value=field)
        ws3.cell(row=row_idx, column=2, value=value)
    _autosize(ws3)

    wb.save(out_path)
    return out_path


def export_summary_only_excel(summary, whole_dataset_max, whole_dataset_max_percent):
    if not config.SAVE_EXCEL:
        return None

    excel_folder = helpers.ensure_dir(os.path.join(config.OUTPUT_FOLDER, "Excel"))
    out_path = os.path.join(excel_folder, "summary_statistics.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary Statistics"
    _write_header(ws, ["Metric", "Mean", "Std Dev", "Min", "Max"])
    _write_summary_rows(ws, summary, whole_dataset_max, whole_dataset_max_percent)
    _autosize(ws)

    wb.save(out_path)
    return out_path
